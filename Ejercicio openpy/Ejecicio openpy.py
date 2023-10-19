import openpyxl

def ingresar_gastos():
    gastos = []
    while True:
        fecha = input("Ingrese la fecha del gasto (YYYY-MM-DD) o 'fin' para salir: ")
        if fecha.lower() == 'fin':
            break
        descripcion = input("Ingrese la descripción del gasto: ")
        monto = float(input("Ingrese el monto del gasto: "))
        gastos.append((fecha, descripcion, monto))
    return gastos

def calcular_resumen(gastos):
    if not gastos:
        return 0, None, None, 0

    total_gastos = sum(gasto[2] for gasto in gastos)
    gasto_mas_caro = max(gastos, key=lambda x: x[2])
    gasto_mas_barato = min(gastos, key=lambda x: x[2])

    return len(gastos), gasto_mas_caro, gasto_mas_barato, total_gastos

def guardar_informe(gastos):
    informe = openpyxl.Workbook()
    hoja = informe.active
    hoja.title = "Gastos"
    hoja['A1'] = "Fecha"
    hoja['B1'] = "Descripción"
    hoja['C1'] = "Monto"

    for i, gasto in enumerate(gastos, start=2):
        hoja.cell(row=i, column=1, value=gasto[0])
        hoja.cell(row=i, column=2, value=gasto[1])
        hoja.cell(row=i, column=3, value=gasto[2])

    informe.save("informe_gastos.xlsx")
    print("El informe de gastos se ha guardado en 'informe_gastos.xlsx'")

def main():
    print("Ingrese los detalles de sus gastos. Para finalizar, escriba 'fin'.")
    gastos = ingresar_gastos()

    num_gastos, gasto_mas_caro, gasto_mas_barato, total_gastos = calcular_resumen(gastos)

    print("\nResumen de gastos:")
    print("Número total de gastos:", num_gastos)
    if num_gastos > 0:
        print("Gasto más caro:")
        print("Fecha:", gasto_mas_caro[0])
        print("Descripción:", gasto_mas_caro[1])
        print("Monto:", gasto_mas_caro[2])
        print("Gasto más barato:")
        print("Fecha:", gasto_mas_barato[0])
        print("Descripción:", gasto_mas_barato[1])
        print("Monto:", gasto_mas_barato[2])
        print("Monto total de gastos:", total_gastos)

    guardar_informe(gastos)

if __name__ == "__main__":
    main()
